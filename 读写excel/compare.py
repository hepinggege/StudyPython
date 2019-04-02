# -*- coding:utf-8 -*-

import sys
import xlwt
import openpyxl


font = openpyxl.styles.Font(color = openpyxl.styles.colors.RED)



base_xls = sys.argv[1]
sec_xls = sys.argv[2]

wb = openpyxl.load_workbook(base_xls)
vb = openpyxl.load_workbook(sec_xls)
# 获取workbook中所有的表格
sheet1 = wb.get_sheet_names()
sheet2 = vb.get_sheet_names()
value1 = []
value2 = []
differ = []

for i in range(len(sheet1)):
    sheet_w = wb.get_sheet_by_name(sheet1[i])
    sheet_v = vb.get_sheet_by_name(sheet2[i])
    for r in range(3, sheet_w.max_row + 1):
        for c in range(2, sheet_w.max_column + 1):
            value1 = sheet_w.cell(row=r,column=c).value
            value2 = sheet_v.cell(row=r, column=c).value
            if (value1 is None) or (value2 is None):
                continue
            a = int(value2) - int(value1)
            print(str(a))
            if (a / value1) >= 0.2:
                sheet_w.cell(row=r, column=c).font = font
                sheet_v.cell(row=r, column=c).font = font
wb.save(base_xls)
vb.save(sec_xls)






