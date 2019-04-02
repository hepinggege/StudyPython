# -*- coding:utf-8 -*-

import sys
import xlwt
import xlrd
import openpyxl
from xlutils.copy import copy

myStyle = xlwt.easyxf('font: name Times New Roman, color-index red, bold on', num_format_str='#,##0.00')


def is_num(s):
    try:
        num = float(s)
        return True, num
    except ValueError:
        return False, 0


base_xls = sys.argv[1]
sec_xls = sys.argv[2]

wb = xlrd.open_workbook(base_xls)
vb = xlrd.open_workbook(sec_xls)
vb_new = copy(vb)
# 获取workbook中所有的表格
sheet1 = wb.sheets()
sheet2 = vb.sheets()

for i in range(len(sheet1)):
    sheet_w = sheet1[i]
    sheet_v = sheet2[i]
    sheet_new = vb_new.get_sheet(i)
    for wRow in range(2, sheet_w.nrows):
        # 先遍历第一个表格的每一行的device name
        wDevice = sheet_w.cell(wRow, 0).value
        for vRow in range(2, sheet_v.nrows):
            vDevice = sheet_v.cell(vRow, 0).value
            # 在第二个表格中遍历找到与之名字相同的device name
            if wDevice == vDevice:
                print("行匹配成功:" + str(wDevice) + ":" + str(vDevice))
                # 找到相同的device name之后， 以相同的逻辑开始匹配列
                for wCol in range(1, sheet_w.ncols):
                    wColName = sheet_w.cell(1, wCol).value
                    for vCol in range(1, sheet_v.ncols):
                        vColName = sheet_v.cell(1, vCol).value
                        if wColName == vColName:
                            print("列匹配成功:" + str(wColName) + ":" + str(vColName))
                            value1 = sheet_w.cell(wRow, wCol).value
                            value2 = sheet_v.cell(vRow, vCol).value
                            isTrue1, num1 = is_num(value1)
                            isTrue2, num2 = is_num(value2)
                            if isTrue1 and isTrue2:
                                a = num2 - num1
                                if num1 != 0 and a / num1 > 0:
                                    sheet_new.write(vRow, vCol, value2, myStyle)
                            break
                        else:
                            print("列匹配失败:" + str(wColName) + ":" + str(vColName))

                # 匹配到了之后结束内循环，开始匹配后面的Device
                break
            else:
                print("行匹配失败:" + str(wDevice) + ":" + str(vDevice))

vb_new.save(sec_xls)







